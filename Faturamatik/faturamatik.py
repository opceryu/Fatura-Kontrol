import openpyxl
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

CorFill = PatternFill(fill_type=None, start_color='ffffff', end_color='21f600')

wb = load_workbook(filename = 'karsilastir.xlsx')
wsRef = wb.get_sheet_by_name('references')
wsComp = wb.get_sheet_by_name('compare')
ErrFont = Font(name='Arial', size=7, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF0000')
CorFont = Font(name='Arial', size=7, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='21f600')

##paydate checker
def checkpaydate(RefDate,RefDate1,RefDate2):
	while RefDate.weekday() != 4:
		RefDate = RefDate + datetime.timedelta(days = 1)
		print(RefDate)
	while RefDate1.weekday() != 4:
		RefDate1 = RefDate1 + datetime.timedelta(days = 1)
	while RefDate2.weekday() != 4:
		RefDate2 = RefDate2 + datetime.timedelta(days = 1)
	if PayDate == RefDate or PayDate == RefDate1 or PayDate == RefDate2:
		return ""
	else:
		return " - Valör Hatalı"

for row in range(3,wsComp.max_row + 1):
#for row in range(3,10):

	##defining variables of comparing entry
	Product = wsComp.cell(row = row, column = 2).value
	Version = str(wsComp.cell(row = row, column = 5).value).partition("-")[0]
	Company = wsComp.cell(row = row, column = 7).value
	PurchaseDate = wsComp.cell(row = row, column = 8).value
	PayDate = wsComp.cell(row = row, column = 10).value
	UnitPrice = wsComp.cell(row = row, column = 13).value
	Amount = wsComp.cell(row = row, column = 12).value
	Revenue = wsComp.cell(row = row, column = 14).value
	Discount = wsComp.cell(row = row, column = 15).value
	NetPrice = wsComp.cell(row = row, column = 16).value
	KDV = wsComp.cell(row = row, column = 17).value
	TotalPrice = wsComp.cell(row = row, column = 18).value
	baseVal = 0
	ErrCell = wsComp.cell(row = row, column = 19)
	ErrCell.font = ErrFont
	Error = ""

	if(Product != None and Version != None and PurchaseDate != None):

		##finding the appropriate reference in the reference section
		for rowR in range(2,wsRef.max_row + 1):
			ProRef = wsRef.cell(row = rowR, column = 1).value
			VerRef = wsRef.cell(row = rowR, column = 2).value
			if Product == ProRef and Version == VerRef:
				refRow = rowR
				break
		else:
			ErrCell.value = "Ürün referanslarda mevcut değil"
			continue

		##findiing the base price for the given date
		for column in range(6, wsRef.max_column):
			dateComp = wsRef.cell(row = 1, column = column).value
			if PurchaseDate <= dateComp:
				baseVal = wsRef.cell(row = refRow, column = column-1).value
				break
		else:
			ErrCell.value = "Tarih yok"
			continue
		if baseVal == None:
			ErrCell.value = "Referans değeri yok"
			continue

		##looking if there is a colakoglu plus value and adding it to the base price
		if wsRef.cell(row = refRow, column = 3).value != None:
			refPrice = float(baseVal) + wsRef.cell(row = refRow, column = 3).value
			colak = True
		else:
			refPrice = baseVal
			colak = False

		##price comparing and checking
		if UnitPrice == refPrice:
			pass	
		else:
			Error = Error + " - Fiyatlar Uyuşmuyor"

		##Checking company to find the correct refere date
		if "BORÇELİK" in Company:
			RD = PurchaseDate + datetime.timedelta(days = wsRef.cell(row = refRow, column = 4).value)
			RD1 = PurchaseDate + datetime.timedelta(days = 1 + wsRef.cell(row = refRow, column = 4).value)
			RD2 = PurchaseDate + datetime.timedelta(days = 2 + wsRef.cell(row = refRow, column = 4).value)
			Error = Error + checkpaydate(RD,RD1,RD2)
			
		elif "BORUSAN" in Company:
			RD = PurchaseDate + datetime.timedelta(days = wsRef.cell(row = refRow, column = 5).value)
			RD1 = PurchaseDate + datetime.timedelta(days = 1 + wsRef.cell(row = refRow, column = 5).value)
			RD2 = PurchaseDate + datetime.timedelta(days = 2 + wsRef.cell(row = refRow, column = 5).value)
			Error = Error + checkpaydate(RD,RD1,RD2)

		else:
			Error = Error + " - Cari Ünvan tanınmıyor; valör kontrol edilemedi"


		##revenue, Net price and KDV+ checking
		if round(UnitPrice * Amount, 2) == Revenue:
			pass
		else:
			Error = Error + " - Tutar hatalı doğrusu = " + str(UnitPrice * Amount)
		if Revenue - Discount == NetPrice:
			pass
		else:
			Error = Error + " - Net fiyat hatalı"
		if NetPrice + KDV == TotalPrice:
			pass
		else:
			Error = Error + " - Toplam Hatalı"

		##writing the Errors on the column
		if Error == "":
			ErrCell.font = CorFont
			ErrCell.value = "Hata yok"		
		else:
			ErrCell.value = Error

	else:
		print("mising info")

wb.save('bitti.xlsx')